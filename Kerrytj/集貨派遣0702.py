from playwright.sync_api import sync_playwright
from getpass import getpass
import clipboard
import openpyxl
import os
import ctypes
ctypes.windll.kernel32.SetConsoleTitleW('集貨派遣')


def read_excel(var, col):
    a = ws1.iter_rows(min_row=0, min_col=col, max_col=col, max_row=ws1.max_row)
    for row in a:
        for cell in row:
            var.append(cell.value)


'''EXCEL讀取'''
wb = openpyxl.load_workbook('發送集貨區.xlsx')
ws1 = wb['集貨區']
roads, gap, min_num, max_num, reign, car, remark = [], [], [], [], [], [], []
read_excel(roads, 3)
read_excel(gap, 4)
read_excel(min_num, 5)
read_excel(max_num, 6)
read_excel(reign, 7)
read_excel(car, 8)
read_excel(remark, 9)

'''初始化'''
print(' 版本：0701 by黃佳葆\n\n\n')
units = ['段', '巷', '弄', '之', '號', '樓']
chi_num = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "零"]
full_num = ["１", "２", "３", "４", "５", "６", "７", "８", "９", "０"]
half_num = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '0']

account = input(" 帳號：")
password = getpass(" 密碼：")
pg = (sync_playwright().start().chromium.launch(headless=False, args=["--start-maximized"])
      .new_context(no_viewport=True).new_page())
pg.goto("http://ktms.kerrytj.com:8088/Portal/EmployeeLogin.jsp")
pg.get_by_placeholder("帳號").fill(account)
pg.get_by_placeholder("帳號").press("Tab")
pg.get_by_placeholder("密碼").fill(password)
pg.get_by_placeholder("密碼").press("Enter")
pg.get_by_text("營業所路線").click()
pg.get_by_text("發送").click()
pg.get_by_text("集貨派遣").click()
with pg.expect_popup() as page1_info:
    pg.get_by_text("集貨派遣處理").click()
last_Page = '0'
last_totalPage = '0'
pg1 = page1_info.value

while True:
    try:
        page = pg1.frame_locator("#fra_DivDialogCS02M0020_iframe").locator("#page").text_content()
        totalPage = pg1.frame_locator("#fra_DivDialogCS02M0020_iframe").locator("#totalPage").text_content()
        if last_Page != page or last_totalPage != totalPage:
            os.system("cls")
            serialNum, address_reign, address_car = 999, 999, 999
            address_road, address_number = "Ｎ", "Ｎ"
            pg1.frame_locator("#fra_DivDialogCS02M0020_iframe").locator("#txtCOLLECTION_ADDR").press("ControlOrMeta+a")
            pg1.frame_locator("#fra_DivDialogCS02M0020_iframe").locator("#txtCOLLECTION_ADDR").press("ControlOrMeta+c")
            address = clipboard.paste()
            address.strip(' ')
            address.strip('　')
            address = address.replace("－", "之")
            address = address.replace("-", "之")
            address = address.replace("f", "樓")
            address = address.replace("F", "樓")
            address = address.replace("Ｆ", "樓")
            address = address.replace("ｆ", "樓")
            address = address.replace("b", "-")
            address = address.replace("B", "-")
            address = address.replace("臺", "台")
            for times in range(10):
                address = address.replace(chi_num[times], half_num[times])
            for times in range(10):
                address = address.replace(full_num[times], half_num[times])
            try:
                '''尋找路名'''
                for i in roads:
                    if address.find(i) != -1:
                        address_road = i
                        serialNum = roads.index(i)
                        address_number = address.split(address_road)[1]
                        break
                '''尋找門牌號'''
                for i in address_number:
                    if i == "段":
                        address_number = address_number.split("段")[0]
                        break
                    elif i == "巷":
                        address_number = address_number.split("巷")[0]
                        break
                    elif i == "弄":
                        address_number = address_number.split("弄")[0]
                        break
                    elif i == "之":
                        address_number = address_number.split("之")[0]
                        break
                    elif i == "號":
                        address_number = address_number.split("號")[0]
                        break
                    elif i == "樓":
                        address_number = address_number.split("樓")[0]
                        break
                '''比對門牌號'''
                while address_road == roads[serialNum]:
                    if gap[serialNum] == 0 and min_num[serialNum] <= int(address_number) <= max_num[serialNum]:
                        address_reign = reign[serialNum]
                        address_car = reign[serialNum]
                        break
                    elif (gap[serialNum] == 1 and min_num[serialNum] <= int(address_number) <= max_num[serialNum]
                          and int(address_number) % 2 == 1):
                        address_reign = reign[serialNum]
                        address_car = reign[serialNum]
                        break
                    elif (gap[serialNum] == 2 and min_num[serialNum] <= int(address_number) <= max_num[serialNum]
                          and int(address_number) % 2 == 0):
                        address_reign = reign[serialNum]
                        address_car = reign[serialNum]
                        break
                    else:
                        serialNum = serialNum + 1
                pg1.frame_locator("#fra_DivDialogCS02M0020_iframe").locator("#txtCOLLECTION_AREA").fill(
                    str(address_car))
                pg1.frame_locator("#fra_DivDialogCS02M0020_iframe").locator("#txtCOLLECTION_AREA").press(
                    "ControlOrMeta+a")
                print(f"\n 地址：{address}")
                print("\n—————————————————————————\n")
                print(f" 路段：{address_road}　｜　號碼：{str(address_number)}")
                print("\n—————————————————————————\n")
                print(f" 集貨區：{str(address_reign)}　｜　車號：{str(address_car)}　｜　備註：{remark[serialNum]}\n\n")
            except:  # 找不到路名或門牌號的情況
                pg1.frame_locator("#fra_DivDialogCS02M0020_iframe").locator("#txtCOLLECTION_AREA").fill(
                    str(address_car))
                pg1.frame_locator("#fra_DivDialogCS02M0020_iframe").locator("#txtCOLLECTION_AREA").press(
                    "ControlOrMeta+a")
                print(f"\n 地址：{address}")
                print("\n—————————————————————————\n")
                print(" 錯誤")
                print("\n—————————————————————————\n")
                print(f" 集貨區：{str(address_reign)}\n\n")
    except:
        page = '0'
        totalPage = '0'

    last_Page = page
    last_totalPage = totalPage
