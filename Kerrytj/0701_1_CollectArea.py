import openpyxl
import os
import ctypes
ctypes.windll.kernel32.SetConsoleTitleW('發送集貨區查詢')


def read_excel(var, col):
    a = ws1.iter_rows(min_row=0, min_col=col, max_col=col, max_row=ws1.max_row)
    for row in a:
        for cell in row:
            var.append(cell.value)


'''EXCEL讀取'''
wb = openpyxl.load_workbook('發送集貨區.xlsx')
ws1 = wb['集貨區']
roads, gap, min_num, max_num, reign, car, remark = [], [], [], [], [], [], []
read_excel(roads, 3)
read_excel(gap, 4)
read_excel(min_num, 5)
read_excel(max_num, 6)
read_excel(reign, 7)
read_excel(car, 8)
read_excel(remark, 9)

'''初始化'''
print(' 版本：0627_no_copy\n\n\n')
chi_num = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "零"]
full_num = ["１", "２", "３", "４", "５", "６", "７", "８", "９", "０"]
half_num = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '0']

while True:  # 開始分析
    serialNum, address_reign, address_car = 999, 999, 999
    address_district, address_road, address_number = "Ｎ", "Ｎ", "Ｎ"
    address = input(' 輸入地址：')
    os.system("cls")
    '''整理地址'''
    address.strip(" ")
    address.strip("　")
    address = address.replace("－", "之")
    address = address.replace("-", "之")
    address = address.replace("f", "樓")
    address = address.replace("F", "樓")
    address = address.replace("Ｆ", "樓")
    address = address.replace("ｆ", "樓")
    address = address.replace("b", "-")
    address = address.replace("B", "-")
    for times in range(10):
        address = address.replace(chi_num[times], half_num[times])
    for times in range(10):
        address = address.replace(full_num[times], half_num[times])

    try:
        '''尋找行政區'''
        districts = ['汐止區', '松山區', '信義區', '中山區', '南港區', '內湖區']
        for i in districts:
            if address.find(i) != -1:
                address_district = i
                break
        '''尋找路名'''
        for i in roads:
            if address.find(i) != -1:
                address_road = i
                serialNum = roads.index(i)
                address_number = address.split(address_road)[1]
                break
        '''尋找門牌號'''
        for i in address_number:
            if i == "段":
                address_number = address_number.split("段")[0]
                break
            elif i == "巷":
                address_number = address_number.split("巷")[0]
                break
            elif i == "弄":
                address_number = address_number.split("弄")[0]
                break
            elif i == "之":
                address_number = address_number.split("之")[0]
                break
            elif i == "號":
                address_number = address_number.split("號")[0]
                break
            elif i == "樓":
                address_number = address_number.split("樓")[0]
                break
        '''比對門牌號'''
        while address_road == roads[serialNum]:
            if gap[serialNum] == 0 and min_num[serialNum] <= int(address_number) <= max_num[serialNum]:
                address_reign = reign[serialNum]
                address_car = reign[serialNum]
                break
            elif gap[serialNum] == 1 and min_num[serialNum] <= int(address_number) <= max_num[serialNum] and int(
                    address_number) % 2 == 1:
                address_reign = reign[serialNum]
                address_car = reign[serialNum]
                break
            elif gap[serialNum] == 2 and min_num[serialNum] <= int(address_number) <= max_num[serialNum] and int(
                    address_number) % 2 == 0:
                address_reign = reign[serialNum]
                address_car = reign[serialNum]
                break
            else:
                serialNum = serialNum + 1
        '''輸出結果'''
        print(f"\n 地址：{address}")
        print("\n—————————————————————————\n")
        print(f" 行政區：{address_district}　｜　路名：{address_road}　｜　號碼：{str(address_number)}")
        print("\n—————————————————————————\n")
        print(f" 集貨區：{str(address_reign)}　｜　車號：{str(address_car)}　｜　備註：{remark[serialNum]}\n\n")
    except:  # 找不到路名或門牌號的情況
        print(f"\n 地址：{address}")
        print("\n—————————————————————————\n")
        print(" 錯誤")
        print("\n—————————————————————————\n")
        print(f" 集貨區：{str(address_reign)}\n\n")
